"""
CSV Import/Export Views for Employees, Public Holidays, and Job Descriptions
"""
import csv
from io import TextIOWrapper
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db import transaction
from datetime import datetime
from decimal import Decimal
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from .models import Employee, Department, Designation, PublicHoliday
from .models_job import JobDescription

#     ====== EMPLOYEE CSV EXPORT/IMPORT     ======

@login_required
def export_employees_csv(request):
    """Export all employees to CSV"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="employees_export.csv"'

    writer = csv.writer(response)

    # Write header
    writer.writerow([
        'Employee Code', 'Full Name', 'Department', 'Designation',
        'Joining Date', 'Relieving Date', 'Employment Status',
        'Mobile Number', 'Official Email', 'Personal Email',
        'Local Address', 'Permanent Address',
        'Date of Birth', 'Marital Status', 'Anniversary Date',
        'Highest Qualification', 'Total Experience Years', 'Total Experience Months',
        'Probation Status', 'Aadhar Card Number', 'PAN Card Number',
        'Emergency Contact Name', 'Emergency Contact Mobile',
        'Emergency Contact Email', 'Emergency Contact Address',
        'Emergency Contact Relationship'
    ])

    # Write data
    employees = Employee.objects.select_related('department', 'designation').all()
    for emp in employees:
        writer.writerow([
            emp.employee_code,
            emp.full_name,
            emp.department.name,
            emp.designation.name,
            emp.joining_date.strftime('%Y-%m-%d') if emp.joining_date else '',
            emp.relieving_date.strftime('%Y-%m-%d') if emp.relieving_date else '',
            emp.employment_status,
            emp.mobile_number,
            emp.official_email,
            emp.personal_email or '',
            emp.local_address or '',
            emp.permanent_address or '',
            emp.date_of_birth.strftime('%Y-%m-%d') if emp.date_of_birth else '',
            emp.marital_status,
            emp.anniversary_date.strftime('%Y-%m-%d') if emp.anniversary_date else '',
            emp.highest_qualification,
            emp.total_experience_years,
            emp.total_experience_months,
            emp.period_type,
            emp.aadhar_card_number,
            emp.pan_card_number,
            emp.emergency_contact_name or '',
            emp.emergency_contact_mobile or '',
            emp.emergency_contact_email or '',
            emp.emergency_contact_address or '',
            emp.emergency_contact_relationship or '',
        ])

    return response

@login_required
def download_employee_sample_csv(request):
    """Download sample CSV template for employee import"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="employee_import_sample.csv"'

    writer = csv.writer(response)

    # Write header
    writer.writerow([
        'Employee Code', 'Full Name', 'Department', 'Designation',
        'Joining Date', 'Relieving Date', 'Employment Status',
        'Mobile Number', 'Official Email', 'Personal Email',
        'Local Address', 'Permanent Address',
        'Date of Birth', 'Marital Status', 'Anniversary Date',
        'Highest Qualification', 'Total Experience Years', 'Total Experience Months',
        'Probation Status', 'Aadhar Card Number', 'PAN Card Number',
        'Emergency Contact Name', 'Emergency Contact Mobile',
        'Emergency Contact Email', 'Emergency Contact Address',
        'Emergency Contact Relationship'
    ])

    # Write sample data
    writer.writerow([
        'EMP001', 'John Doe', 'Engineering', 'Software Engineer',
        '2024-01-15', '', 'active',
        '9876543210', 'john.doe@company.com', 'john@personal.com',
        '123 Street, City', '456 Avenue, Town',
        '1990-05-15', 'married', '2015-06-20',
        'B.Tech Computer Science', '5', '6',
        'Confirmed', '123456789012', 'ABCDE1234F',
        'Jane Doe', '9876543211', 'jane@email.com', '789 Road, Village', 'Spouse'
    ])

    return response

@login_required
def import_employees_csv(request):
    """Import employees from CSV file"""
    if request.method == 'POST' and request.FILES.get('csv_file'):
        csv_file = request.FILES['csv_file']

        if not csv_file.name.endswith('.csv'):
            messages.error(request, 'Please upload a valid CSV file.')
            return redirect('employees:employee_list')

        try:
            # Read CSV file
            file_data = TextIOWrapper(csv_file.file, encoding='utf-8')
            csv_reader = csv.DictReader(file_data)

            success_count = 0
            error_count = 0
            errors = []

            with transaction.atomic():
                for row_num, row in enumerate(csv_reader, start=2):
                    try:
                        # Get or create department
                        dept_name = row.get('Department', '').strip()
                        if not dept_name:
                            raise ValueError("Department is required")

                        department, _ = Department.objects.get_or_create(
                            name=dept_name,
                            defaults={'description': f'{dept_name} Department'}
                        )

                        # Get or create designation
                        desig_name = row.get('Designation', '').strip()
                        if not desig_name:
                            raise ValueError("Designation is required")

                        designation, _ = Designation.objects.get_or_create(
                            name=desig_name,
                            department=department,
                            defaults={'description': f'{desig_name} Position'}
                        )

                        # Parse dates
                        joining_date = datetime.strptime(row.get('Joining Date', '').strip(), '%Y-%m-%d').date() if row.get('Joining Date', '').strip() else None
                        relieving_date = datetime.strptime(row.get('Relieving Date', '').strip(), '%Y-%m-%d').date() if row.get('Relieving Date', '').strip() else None
                        date_of_birth = datetime.strptime(row.get('Date of Birth', '').strip(), '%Y-%m-%d').date() if row.get('Date of Birth', '').strip() else None
                        anniversary_date = datetime.strptime(row.get('Anniversary Date', '').strip(), '%Y-%m-%d').date() if row.get('Anniversary Date', '').strip() else None

                        # Create or update employee
                        employee_code = row.get('Employee Code', '').strip()
                        official_email = row.get('Official Email', '').strip()

                        if not official_email:
                            raise ValueError("Official Email is required")

                        employee, created = Employee.objects.update_or_create(
                            official_email=official_email,
                            defaults={
                                'employee_code': employee_code,
                                'full_name': row.get('Full Name', '').strip(),
                                'department': department,
                                'designation': designation,
                                'joining_date': joining_date,
                                'relieving_date': relieving_date,
                                'employment_status': row.get('Employment Status', 'active').strip(),
                                'mobile_number': row.get('Mobile Number', '').strip(),
                                'personal_email': row.get('Personal Email', '').strip() or None,
                                'local_address': row.get('Local Address', '').strip(),
                                'permanent_address': row.get('Permanent Address', '').strip(),
                                'date_of_birth': date_of_birth,
                                'marital_status': row.get('Marital Status', 'single').strip(),
                                'anniversary_date': anniversary_date,
                                'highest_qualification': row.get('Highest Qualification', '').strip(),
                                'total_experience_years': int(row.get('Total Experience Years', 0) or 0),
                                'total_experience_months': int(row.get('Total Experience Months', 0) or 0),
                                'period_type': row.get('Period Type', 'confirmed').strip(),
                                'aadhar_card_number': row.get('Aadhar Card Number', '').strip(),
                                'pan_card_number': row.get('PAN Card Number', '').strip().upper(),
                                'emergency_contact_name': row.get('Emergency Contact Name', '').strip(),
                                'emergency_contact_mobile': row.get('Emergency Contact Mobile', '').strip(),
                                'emergency_contact_email': row.get('Emergency Contact Email', '').strip(),
                                'emergency_contact_address': row.get('Emergency Contact Address', '').strip(),
                                'emergency_contact_relationship': row.get('Emergency Contact Relationship', '').strip(),
                            }
                        )

                        success_count += 1

                    except Exception as e:
                        error_count += 1
                        errors.append(f"Row {row_num}: {str(e)}")

            # Show results
            if success_count > 0:
                messages.success(request, f'Successfully imported {success_count} employee(s).')

            if error_count > 0:
                error_msg = f'Failed to import {error_count} employee(s). Errors: ' + '; '.join(errors[:5])
                if len(errors) > 5:
                    error_msg += f' ... and {len(errors) - 5} more errors.'
                messages.error(request, error_msg)

        except Exception as e:
            messages.error(request, f'Error processing CSV file: {str(e)}')

        return redirect('employees:employee_list')

    return render(request, 'employees/import_csv.html', {'model_name': 'Employee'})

#     ====== PUBLIC HOLIDAY CSV EXPORT/IMPORT     ======

@login_required
def export_public_holidays_csv(request):
    """Export public holidays to Excel with separate sheets for each country"""
    from django.utils import timezone
    
    # Create a new Excel workbook
    wb = openpyxl.Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Define styles
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='332666', end_color='332666', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    # Get all countries from the model choices
    countries = PublicHoliday.COUNTRY_CHOICES
    
    # Calculate year range from all holidays
    all_holidays = PublicHoliday.objects.all()
    if all_holidays.exists():
        years = sorted(set(holiday.year for holiday in all_holidays))
        year_range = f"{years[0]}-{years[-1]}" if len(years) > 1 else str(years[0])
    else:
        year_range = str(timezone.now().year)
    
    for country_code, country_name in countries:
        # Create sheet for each country
        sheet_name = f"{country_name} Holidays"
        ws = wb.create_sheet(title=sheet_name)
        
        # Write headers
        headers = ['Holiday Name', 'Date', 'Day', 'Year', 'Description', 'Is Optional', 'Is Active']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Get holidays for this country
        holidays = PublicHoliday.objects.filter(country=country_code).order_by('date')
        
        # Write data
        for row_num, holiday in enumerate(holidays, 2):
            ws.cell(row=row_num, column=1, value=holiday.name)
            ws.cell(row=row_num, column=2, value=holiday.date.strftime('%Y-%m-%d'))
            ws.cell(row=row_num, column=3, value=holiday.day)
            ws.cell(row=row_num, column=4, value=holiday.year)
            ws.cell(row=row_num, column=5, value=holiday.description or '')
            ws.cell(row=row_num, column=6, value='Yes' if holiday.is_optional else 'No')
            ws.cell(row=row_num, column=7, value='Yes' if holiday.is_active else 'No')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add summary at the top if there are holidays
        if holidays.count() > 0:
            # Calculate year range for this country
            country_years = sorted(set(holiday.year for holiday in holidays))
            country_year_range = f"{country_years[0]}-{country_years[-1]}" if len(country_years) > 1 else str(country_years[0])
            
            # Insert title row
            ws.insert_rows(1)
            ws.cell(row=1, column=1, value=f"{country_name} Public Holidays - {country_year_range}")
            ws.cell(row=1, column=1).font = Font(bold=True, size=14)
            ws.merge_cells(f'A1:{get_column_letter(len(headers))}1')
            ws.cell(row=1, column=1).alignment = Alignment(horizontal='center')
            
            # Insert total count row
            ws.insert_rows(2)
            ws.cell(row=2, column=1, value=f"Total Holidays: {holidays.count()}")
            ws.cell(row=2, column=1).font = Font(bold=True)
            ws.merge_cells(f'A2:{get_column_letter(len(headers))}2')
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'public_holidays_all_countries_{year_range}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Save workbook to response
    wb.save(response)
    
    return response

@login_required
def download_public_holiday_sample_csv(request):
    """Download sample CSV template for public holiday import"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="public_holiday_import_sample.csv"'

    writer = csv.writer(response)

    # Write header
    writer.writerow(['Name', 'Date', 'Year', 'Is Active'])

    # Write sample data
    writer.writerow(['Republic Day', '2026-01-26', '2026', 'Yes'])
    writer.writerow(['Independence Day', '2026-08-15', '2026', 'Yes'])
    writer.writerow(['Gandhi Jayanti', '2026-10-02', '2026', 'Yes'])
    writer.writerow(['Diwali', '2026-11-04', '2026', 'Yes'])
    writer.writerow(['Christmas', '2026-12-25', '2026', 'Yes'])

    return response

@login_required
def import_public_holidays_csv(request):
    """Import public holidays from CSV file"""
    if request.method == 'POST' and request.FILES.get('csv_file'):
        csv_file = request.FILES['csv_file']

        if not csv_file.name.endswith('.csv'):
            messages.error(request, 'Please upload a valid CSV file.')
            return redirect('employees:public_holidays')

        try:
            # Read CSV file
            file_data = TextIOWrapper(csv_file.file, encoding='utf-8')
            csv_reader = csv.DictReader(file_data)

            success_count = 0
            error_count = 0
            errors = []

            with transaction.atomic():
                for row_num, row in enumerate(csv_reader, start=2):
                    try:
                        # Parse date
                        date_str = row.get('Date', '').strip()
                        if not date_str:
                            raise ValueError("Date is required")

                        holiday_date = datetime.strptime(date_str, '%Y-%m-%d').date()

                        # Get day name from date
                        day_name = holiday_date.strftime('%A')

                        # Parse is_active
                        is_active_str = row.get('Is Active', 'Yes').strip().lower()
                        is_active = is_active_str in ['yes', 'true', '1', 'active']

                        # Create or update holiday
                        holiday, created = PublicHoliday.objects.update_or_create(
                            date=holiday_date,
                            defaults={
                                'name': row.get('Name', '').strip(),
                                'day': day_name,
                                'year': int(row.get('Year', holiday_date.year)),
                                'is_active': is_active
                            }
                        )

                        success_count += 1

                    except Exception as e:
                        error_count += 1
                        errors.append(f"Row {row_num}: {str(e)}")

            # Show results
            if success_count > 0:
                messages.success(request, f'Successfully imported {success_count} public holiday(s).')

            if error_count > 0:
                error_msg = f'Failed to import {error_count} holiday(s). Errors: ' + '; '.join(errors[:5])
                if len(errors) > 5:
                    error_msg += f' ... and {len(errors) - 5} more errors.'
                messages.error(request, error_msg)

        except Exception as e:
            messages.error(request, f'Error processing CSV file: {str(e)}')

        return redirect('employees:public_holidays')

    return render(request, 'employees/import_csv.html', {'model_name': 'Public Holiday'})

#     ====== JOB DESCRIPTION CSV EXPORT/IMPORT     ======

@login_required
def export_job_descriptions_csv(request):
    """Export all job descriptions to CSV"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="job_descriptions_export.csv"'

    writer = csv.writer(response)

    # Write header
    writer.writerow([
        'Title', 'Department', 'Designation', 'Employment Type', 'Experience Level',
        'Required Qualifications', 'Skills & Requirements',
        'Min Salary', 'Max Salary', 'Currency', 'Location', 'Work Mode',
        'Number of Vacancies', 'Application Deadline', 'Status'
    ])

    # Write data
    jobs = JobDescription.objects.select_related('department', 'designation').all()
    for job in jobs:
        writer.writerow([
            job.title,
            job.department.name,
            job.designation.name,
            job.employment_type,
            job.experience_level,
            job.required_qualifications,
            job.skills_requirements,
            str(job.min_salary) if job.min_salary else '',
            str(job.max_salary) if job.max_salary else '',
            job.currency,
            job.location,
            job.work_mode,
            job.number_of_vacancies,
            job.application_deadline.strftime('%Y-%m-%d') if job.application_deadline else '',
            job.status
        ])

    return response

@login_required
def download_job_description_sample_csv(request):
    """Download sample CSV template for job description import"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="job_description_import_sample.csv"'

    writer = csv.writer(response)

    # Write header
    writer.writerow([
        'Title', 'Department', 'Designation', 'Employment Type', 'Experience Level',
        'Required Qualifications', 'Skills & Requirements',
        'Min Salary', 'Max Salary', 'Currency', 'Location', 'Work Mode',
        'Number of Vacancies', 'Application Deadline', 'Status'
    ])

    # Write sample data
    writer.writerow([
        'Senior Software Engineer',
        'Engineering',
        'Software Engineer',
        'full_time',
        'senior',
        'Bachelor degree in Computer Science; 5+ years experience in software development',
        'Python, Django, React, PostgreSQL, REST APIs, Git',
        '800000',
        '1200000',
        'INR',
        'Bangalore',
        'Hybrid',
        '2',
        '2026-03-31',
        'active'
    ])

    return response

@login_required
def import_job_descriptions_csv(request):
    """Import job descriptions from CSV file"""
    if request.method == 'POST' and request.FILES.get('csv_file'):
        csv_file = request.FILES['csv_file']

        if not csv_file.name.endswith('.csv'):
            messages.error(request, 'Please upload a valid CSV file.')
            return redirect('employees:job_list')

        try:
            # Read CSV file
            file_data = TextIOWrapper(csv_file.file, encoding='utf-8')
            csv_reader = csv.DictReader(file_data)

            success_count = 0
            error_count = 0
            errors = []

            with transaction.atomic():
                for row_num, row in enumerate(csv_reader, start=2):
                    try:
                        # Get or create department
                        dept_name = row.get('Department', '').strip()
                        if not dept_name:
                            raise ValueError("Department is required")

                        department, _ = Department.objects.get_or_create(
                            name=dept_name,
                            defaults={'description': f'{dept_name} Department'}
                        )

                        # Get or create designation
                        desig_name = row.get('Designation', '').strip()
                        if not desig_name:
                            raise ValueError("Designation is required")

                        designation, _ = Designation.objects.get_or_create(
                            name=desig_name,
                            department=department,
                            defaults={'description': f'{desig_name} Position'}
                        )

                        # Parse dates
                        application_deadline = None
                        deadline_str = row.get('Application Deadline', '').strip()
                        if deadline_str:
                            application_deadline = datetime.strptime(deadline_str, '%Y-%m-%d').date()

                        # Parse boolean fields

                        # Parse salary
                        min_salary = None
                        max_salary = None
                        if row.get('Min Salary', '').strip():
                            min_salary = Decimal(row.get('Min Salary', '').strip())
                        if row.get('Max Salary', '').strip():
                            max_salary = Decimal(row.get('Max Salary', '').strip())

                        # Create job description
                        job = JobDescription.objects.create(
                            title=row.get('Title', '').strip(),
                            department=department,
                            designation=designation,
                            employment_type=row.get('Employment Type', 'full_time').strip(),
                            experience_level=row.get('Experience Level', 'fresher').strip(),
                            required_qualifications=row.get('Required Qualifications', '').strip(),
                            skills_requirements=row.get('Skills & Requirements', '').strip(),
                            min_salary=min_salary,
                            max_salary=max_salary,
                            currency=row.get('Currency', 'INR').strip(),
                            location=row.get('Location', '').strip(),
                            work_mode=row.get('Work Mode', 'Office').strip(),
                            number_of_vacancies=int(row.get('Number of Vacancies', 1) or 1),
                            application_deadline=application_deadline,
                            status=row.get('Status', 'draft').strip(),
                            posted_by=request.user
                        )

                        success_count += 1

                    except Exception as e:
                        error_count += 1
                        errors.append(f"Row {row_num}: {str(e)}")

            # Show results
            if success_count > 0:
                messages.success(request, f'Successfully imported {success_count} job description(s).')

            if error_count > 0:
                error_msg = f'Failed to import {error_count} job(s). Errors: ' + '; '.join(errors[:5])
                if len(errors) > 5:
                    error_msg += f' ... and {len(errors) - 5} more errors.'
                messages.error(request, error_msg)

        except Exception as e:
            messages.error(request, f'Error processing CSV file: {str(e)}')

        return redirect('employees:job_list')

    return render(request, 'employees/import_csv.html', {'model_name': 'Job Description'})
